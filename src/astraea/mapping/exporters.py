"""Export DomainMappingSpec to Excel and JSON formats.

Provides two export functions:
- export_to_json: Pydantic serialization to JSON file
- export_to_excel: openpyxl workbook with 3 sheets (Mapping Spec, Unmapped Variables, Summary)
  including conditional formatting for confidence levels.
"""

from __future__ import annotations

from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from astraea.models.mapping import DomainMappingSpec


def export_to_json(spec: DomainMappingSpec, output_path: Path) -> Path:
    """Export a mapping specification to JSON via Pydantic serialization.

    Args:
        spec: The domain mapping specification to export.
        output_path: File path to write the JSON output.

    Returns:
        The path written.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(spec.model_dump_json(indent=2))
    logger.info("Exported mapping spec to JSON: {path}", path=output_path)
    return output_path


def export_to_excel(spec: DomainMappingSpec, output_path: Path) -> Path:
    """Export a mapping specification to an Excel workbook with 3 sheets.

    Sheet 1 - Mapping Spec: All variable mappings with conditional formatting
        on the Confidence Level column (GREEN=HIGH, YELLOW=MEDIUM, RED=LOW).
    Sheet 2 - Unmapped Variables: Source variables not mapped, plus SUPPQUAL candidates.
    Sheet 3 - Summary: Domain metadata and mapping statistics.

    Args:
        spec: The domain mapping specification to export.
        output_path: File path to write the .xlsx output.

    Returns:
        The path written.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Mapping Spec ---
    ws_mapping = wb.active
    ws_mapping.title = "Mapping Spec"  # type: ignore[union-attr]
    _write_mapping_sheet(ws_mapping, spec)  # type: ignore[arg-type]

    # --- Sheet 2: Unmapped Variables ---
    ws_unmapped = wb.create_sheet("Unmapped Variables")
    _write_unmapped_sheet(ws_unmapped, spec)

    # --- Sheet 3: Summary ---
    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, spec)

    wb.save(output_path)
    logger.info("Exported mapping spec to Excel: {path}", path=output_path)
    return output_path


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_MAPPING_HEADERS = [
    "Row #",
    "SDTM Variable",
    "SDTM Label",
    "Type",
    "Core",
    "Source Dataset",
    "Source Variable",
    "Source Label",
    "Pattern",
    "Mapping Logic",
    "Derivation Rule",
    "CT Codelist",
    "Confidence",
    "Confidence Level",
    "Notes",
]

_HEADER_FONT = Font(bold=True)
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Reasonable default column widths
_COL_WIDTHS = {
    "Row #": 7,
    "SDTM Variable": 16,
    "SDTM Label": 35,
    "Type": 6,
    "Core": 6,
    "Source Dataset": 18,
    "Source Variable": 16,
    "Source Label": 30,
    "Pattern": 14,
    "Mapping Logic": 40,
    "Derivation Rule": 30,
    "CT Codelist": 12,
    "Confidence": 11,
    "Confidence Level": 16,
    "Notes": 30,
}


def _write_mapping_sheet(ws: object, spec: DomainMappingSpec) -> None:
    """Populate the Mapping Spec sheet."""
    # Write headers
    for col_idx, header in enumerate(_MAPPING_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
        cell.font = _HEADER_FONT

    # Set column widths
    for col_idx, header in enumerate(_MAPPING_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _COL_WIDTHS.get(header, 15)  # type: ignore[union-attr]

    # Write data rows
    for row_idx, mapping in enumerate(spec.variable_mappings, start=1):
        data_row = row_idx + 1  # +1 for header
        ws.cell(row=data_row, column=1, value=row_idx)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=2, value=mapping.sdtm_variable)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=3, value=mapping.sdtm_label)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=4, value=mapping.sdtm_data_type)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=5, value=mapping.core.value)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=6, value=mapping.source_dataset or "")  # type: ignore[union-attr]
        ws.cell(row=data_row, column=7, value=mapping.source_variable or "")  # type: ignore[union-attr]
        ws.cell(row=data_row, column=8, value=mapping.source_label or "")  # type: ignore[union-attr]
        ws.cell(row=data_row, column=9, value=mapping.mapping_pattern.value)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=10, value=mapping.mapping_logic)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=11, value=mapping.derivation_rule or "")  # type: ignore[union-attr]
        ws.cell(row=data_row, column=12, value=mapping.codelist_code or "")  # type: ignore[union-attr]
        ws.cell(row=data_row, column=13, value=mapping.confidence)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=14, value=mapping.confidence_level.value)  # type: ignore[union-attr]
        ws.cell(row=data_row, column=15, value=mapping.notes)  # type: ignore[union-attr]

    # Auto-filter on header row
    if spec.variable_mappings:
        last_col = get_column_letter(len(_MAPPING_HEADERS))
        last_row = len(spec.variable_mappings) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"  # type: ignore[union-attr]

    # Conditional formatting on "Confidence Level" column (column 14 = N)
    conf_col = get_column_letter(14)
    if spec.variable_mappings:
        range_str = f"{conf_col}2:{conf_col}{len(spec.variable_mappings) + 1}"

        ws.conditional_formatting.add(  # type: ignore[union-attr]
            range_str,
            CellIsRule(
                operator="equal",
                formula=['"high"'],
                fill=_GREEN_FILL,
            ),
        )
        ws.conditional_formatting.add(  # type: ignore[union-attr]
            range_str,
            CellIsRule(
                operator="equal",
                formula=['"medium"'],
                fill=_YELLOW_FILL,
            ),
        )
        ws.conditional_formatting.add(  # type: ignore[union-attr]
            range_str,
            CellIsRule(
                operator="equal",
                formula=['"low"'],
                fill=_RED_FILL,
            ),
        )


def _write_unmapped_sheet(ws: object, spec: DomainMappingSpec) -> None:
    """Populate the Unmapped Variables sheet."""
    headers = ["Source Dataset", "Source Variable", "Source Label", "Disposition"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
        cell.font = _HEADER_FONT

    # Set column widths
    widths = [18, 18, 35, 18]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]

    row = 2

    # Unmapped source variables
    for var_name in spec.unmapped_source_variables:
        ws.cell(row=row, column=1, value=spec.source_datasets[0] if spec.source_datasets else "")  # type: ignore[union-attr]
        ws.cell(row=row, column=2, value=var_name)  # type: ignore[union-attr]
        ws.cell(row=row, column=3, value="")  # type: ignore[union-attr]
        ws.cell(row=row, column=4, value="Unmapped")  # type: ignore[union-attr]
        row += 1

    # SUPPQUAL candidates
    for var_name in spec.suppqual_candidates:
        ws.cell(row=row, column=1, value=spec.source_datasets[0] if spec.source_datasets else "")  # type: ignore[union-attr]
        ws.cell(row=row, column=2, value=var_name)  # type: ignore[union-attr]
        ws.cell(row=row, column=3, value="")  # type: ignore[union-attr]
        ws.cell(row=row, column=4, value="SUPPDM Candidate")  # type: ignore[union-attr]
        row += 1


def _write_summary_sheet(ws: object, spec: DomainMappingSpec) -> None:
    """Populate the Summary sheet."""
    label_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 25  # type: ignore[union-attr]
    ws.column_dimensions["B"].width = 50  # type: ignore[union-attr]

    rows: list[tuple[str, str | int | float]] = [
        ("Domain", spec.domain),
        ("Domain Label", spec.domain_label),
        ("Domain Class", spec.domain_class),
        ("Structure", spec.structure),
        ("Study ID", spec.study_id),
        ("Mapping Timestamp", spec.mapping_timestamp),
        ("Model Used", spec.model_used),
        ("", ""),
        ("Total Variables Mapped", spec.total_variables),
        ("Required Mapped", spec.required_mapped),
        ("Expected Mapped", spec.expected_mapped),
        ("HIGH Confidence", spec.high_confidence_count),
        ("MEDIUM Confidence", spec.medium_confidence_count),
        ("LOW Confidence", spec.low_confidence_count),
        ("", ""),
        ("Source Datasets", ", ".join(spec.source_datasets)),
        (
            "Cross-Domain Sources",
            ", ".join(spec.cross_domain_sources) if spec.cross_domain_sources else "None",
        ),
        ("Unmapped Variables", len(spec.unmapped_source_variables)),
        ("SUPPQUAL Candidates", len(spec.suppqual_candidates)),
    ]

    for row_idx, (label, value) in enumerate(rows, start=1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)  # type: ignore[union-attr]
        label_cell.font = label_font
        value_cell = ws.cell(row=row_idx, column=2, value=value)  # type: ignore[union-attr]
        value_cell.alignment = wrap_align
