"""Tests for mapping spec exporters (JSON and Excel).

Verifies that DomainMappingSpec can be exported to JSON with round-trip
fidelity, and to Excel with 3 sheets, correct headers, data rows,
conditional formatting, and summary values.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from astraea.mapping.exporters import export_to_excel, export_to_json
from astraea.models.mapping import (
    ConfidenceLevel,
    DomainMappingSpec,
    MappingPattern,
    VariableMapping,
)
from astraea.models.sdtm import CoreDesignation


@pytest.fixture()
def sample_spec() -> DomainMappingSpec:
    """Realistic DomainMappingSpec with mixed confidence levels."""
    mappings = [
        VariableMapping(
            sdtm_variable="STUDYID",
            sdtm_label="Study Identifier",
            sdtm_data_type="Char",
            core=CoreDesignation.REQ,
            source_dataset=None,
            source_variable=None,
            source_label=None,
            mapping_pattern=MappingPattern.ASSIGN,
            mapping_logic="Assign constant study ID",
            assigned_value="PHA022121-C301",
            confidence=0.95,
            confidence_level=ConfidenceLevel.HIGH,
            confidence_rationale="Standard assignment",
            notes="",
        ),
        VariableMapping(
            sdtm_variable="DOMAIN",
            sdtm_label="Domain Abbreviation",
            sdtm_data_type="Char",
            core=CoreDesignation.REQ,
            source_dataset=None,
            source_variable=None,
            source_label=None,
            mapping_pattern=MappingPattern.ASSIGN,
            mapping_logic="Assign 'DM'",
            assigned_value="DM",
            confidence=1.0,
            confidence_level=ConfidenceLevel.HIGH,
            confidence_rationale="Always DM",
            notes="",
        ),
        VariableMapping(
            sdtm_variable="USUBJID",
            sdtm_label="Unique Subject Identifier",
            sdtm_data_type="Char",
            core=CoreDesignation.REQ,
            source_dataset="dm.sas7bdat",
            source_variable="Subject",
            source_label="Subject Number",
            mapping_pattern=MappingPattern.DERIVATION,
            mapping_logic="STUDYID || '-' || SiteNumber || '-' || Subject",
            derivation_rule="CONCAT(STUDYID, '-', SiteNumber, '-', Subject)",
            confidence=0.90,
            confidence_level=ConfidenceLevel.HIGH,
            confidence_rationale="Standard derivation",
            notes="",
        ),
        VariableMapping(
            sdtm_variable="SEX",
            sdtm_label="Sex",
            sdtm_data_type="Char",
            core=CoreDesignation.REQ,
            source_dataset="dm.sas7bdat",
            source_variable="SEX",
            source_label="Gender",
            mapping_pattern=MappingPattern.LOOKUP_RECODE,
            mapping_logic="Map via CT codelist C66731",
            codelist_code="C66731",
            codelist_name="Sex",
            confidence=0.70,
            confidence_level=ConfidenceLevel.MEDIUM,
            confidence_rationale="CT lookup needed",
            notes="Verify coded values",
        ),
        VariableMapping(
            sdtm_variable="RACE",
            sdtm_label="Race",
            sdtm_data_type="Char",
            core=CoreDesignation.EXP,
            source_dataset="dm.sas7bdat",
            source_variable="RACE",
            source_label="Race",
            mapping_pattern=MappingPattern.LOOKUP_RECODE,
            mapping_logic="Map via CT codelist C74457",
            codelist_code="C74457",
            codelist_name="Race",
            confidence=0.50,
            confidence_level=ConfidenceLevel.LOW,
            confidence_rationale="Multiple values possible",
            notes="Review free text",
        ),
        VariableMapping(
            sdtm_variable="BRTHDTC",
            sdtm_label="Date/Time of Birth",
            sdtm_data_type="Char",
            core=CoreDesignation.EXP,
            source_dataset="dm.sas7bdat",
            source_variable="BRTHDAT",
            source_label="Date of Birth",
            mapping_pattern=MappingPattern.REFORMAT,
            mapping_logic="Convert SAS date to ISO 8601",
            confidence=0.85,
            confidence_level=ConfidenceLevel.HIGH,
            confidence_rationale="Standard date conversion",
            notes="",
        ),
    ]

    return DomainMappingSpec(
        domain="DM",
        domain_label="Demographics",
        domain_class="Special Purpose",
        structure="One record per subject",
        study_id="PHA022121-C301",
        source_datasets=["dm.sas7bdat"],
        cross_domain_sources=["ex.sas7bdat"],
        variable_mappings=mappings,
        total_variables=6,
        required_mapped=4,
        expected_mapped=2,
        high_confidence_count=4,
        medium_confidence_count=1,
        low_confidence_count=1,
        mapping_timestamp="2026-02-27T12:00:00Z",
        model_used="claude-sonnet-4-20250514",
        unmapped_source_variables=["SCREENED", "ENROLLED"],
        suppqual_candidates=["ETHNGRP"],
    )


# -----------------------------------------------------------------------
# JSON export tests
# -----------------------------------------------------------------------


class TestExportToJSON:
    """Tests for JSON export and round-trip fidelity."""

    def test_export_to_json(self, sample_spec: DomainMappingSpec, tmp_path: Path) -> None:
        """JSON export creates a valid file that round-trips back to DomainMappingSpec."""
        out = tmp_path / "dm_mapping.json"
        result = export_to_json(sample_spec, out)

        assert result == out
        assert out.exists()

        # Round-trip
        restored = DomainMappingSpec.model_validate_json(out.read_text())
        assert restored.domain == sample_spec.domain
        assert len(restored.variable_mappings) == len(sample_spec.variable_mappings)
        assert restored.study_id == sample_spec.study_id
        assert restored.total_variables == sample_spec.total_variables

    def test_export_to_json_creates_parent_dirs(
        self, sample_spec: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """JSON export creates parent directories if they don't exist."""
        out = tmp_path / "nested" / "dirs" / "dm_mapping.json"
        export_to_json(sample_spec, out)
        assert out.exists()


# -----------------------------------------------------------------------
# Excel export tests
# -----------------------------------------------------------------------


class TestExportToExcel:
    """Tests for Excel export with 3 sheets, formatting, and data."""

    def test_export_to_excel(self, sample_spec: DomainMappingSpec, tmp_path: Path) -> None:
        """Excel export creates a file with 3 sheets and correct header/data rows."""
        out = tmp_path / "dm_mapping.xlsx"
        result = export_to_excel(sample_spec, out)

        assert result == out
        assert out.exists()

        wb = load_workbook(out)
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames == ["Mapping Spec", "Unmapped Variables", "Summary"]

        # Check Mapping Spec sheet
        ws = wb["Mapping Spec"]
        # Header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert headers[0] == "Row #"
        assert headers[1] == "SDTM Variable"
        assert headers[13] == "Confidence Level"

        # Data row count (6 mappings = rows 2-7)
        assert ws.cell(row=2, column=2).value == "STUDYID"
        assert ws.cell(row=7, column=2).value == "BRTHDTC"
        assert ws.cell(row=8, column=2).value is None  # No extra rows

    def test_export_to_excel_conditional_formatting(
        self, sample_spec: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """Confidence Level column has conditional formatting rules for HIGH/MEDIUM/LOW."""
        out = tmp_path / "dm_mapping.xlsx"
        export_to_excel(sample_spec, out)

        wb = load_workbook(out)
        ws = wb["Mapping Spec"]

        # Conditional formatting should exist on the sheet
        cf_rules = ws.conditional_formatting
        assert len(list(cf_rules)) > 0

        # Collect all rules across all ranges
        formulas_found: list[str] = []
        for cf in cf_rules:
            for rule in cf.rules:
                if hasattr(rule, "formula") and rule.formula:
                    formulas_found.extend(rule.formula)

        # Verify all three confidence levels have rules
        formula_text = " ".join(formulas_found)
        assert '"high"' in formula_text
        assert '"medium"' in formula_text
        assert '"low"' in formula_text

    def test_export_to_excel_unmapped_sheet(
        self, sample_spec: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """Unmapped Variables sheet shows unmapped vars and SUPPQUAL candidates."""
        out = tmp_path / "dm_mapping.xlsx"
        export_to_excel(sample_spec, out)

        wb = load_workbook(out)
        ws = wb["Unmapped Variables"]

        # Headers
        assert ws.cell(row=1, column=1).value == "Source Dataset"
        assert ws.cell(row=1, column=4).value == "Disposition"

        # Unmapped: SCREENED (row 2), ENROLLED (row 3)
        assert ws.cell(row=2, column=2).value == "SCREENED"
        assert ws.cell(row=2, column=4).value == "Unmapped"
        assert ws.cell(row=3, column=2).value == "ENROLLED"
        assert ws.cell(row=3, column=4).value == "Unmapped"

        # SUPPQUAL: ETHNGRP (row 4)
        assert ws.cell(row=4, column=2).value == "ETHNGRP"
        assert ws.cell(row=4, column=4).value == "SUPPDM Candidate"

    def test_export_to_excel_summary_sheet(
        self, sample_spec: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """Summary sheet has correct metadata and statistics."""
        out = tmp_path / "dm_mapping.xlsx"
        export_to_excel(sample_spec, out)

        wb = load_workbook(out)
        ws = wb["Summary"]

        # Build a dict from the summary rows
        summary: dict[str, object] = {}
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            if label:
                summary[label] = value

        assert summary["Domain"] == "DM"
        assert summary["Domain Label"] == "Demographics"
        assert summary["Study ID"] == "PHA022121-C301"
        assert summary["Total Variables Mapped"] == 6
        assert summary["Required Mapped"] == 4
        assert summary["Expected Mapped"] == 2
        assert summary["HIGH Confidence"] == 4
        assert summary["MEDIUM Confidence"] == 1
        assert summary["LOW Confidence"] == 1
        assert summary["Source Datasets"] == "dm.sas7bdat"

    def test_export_to_excel_suppqual_label_dynamic_domain(self, tmp_path: Path) -> None:
        """SUPPQUAL candidate label uses dynamic domain name, not hardcoded 'DM'."""
        ae_spec = DomainMappingSpec(
            domain="AE",
            domain_label="Adverse Events",
            domain_class="Events",
            structure="One record per AE per subject",
            study_id="PHA022121-C301",
            source_datasets=["ae.sas7bdat"],
            variable_mappings=[],
            total_variables=0,
            required_mapped=0,
            expected_mapped=0,
            high_confidence_count=0,
            medium_confidence_count=0,
            low_confidence_count=0,
            mapping_timestamp="2026-02-27T12:00:00Z",
            model_used="claude-sonnet-4-20250514",
            unmapped_source_variables=[],
            suppqual_candidates=["CUSTOM_VAR"],
        )

        out = tmp_path / "ae_mapping.xlsx"
        export_to_excel(ae_spec, out)

        wb = load_workbook(out)
        ws = wb["Unmapped Variables"]

        # SUPPQUAL candidate should say "SUPPAE Candidate", not "SUPPDM Candidate"
        assert ws.cell(row=2, column=2).value == "CUSTOM_VAR"
        assert ws.cell(row=2, column=4).value == "SUPPAE Candidate"

    def test_export_to_excel_suppqual_label_dm_domain(
        self, sample_spec: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """SUPPQUAL label for DM domain still says 'SUPPDM Candidate'."""
        out = tmp_path / "dm_mapping.xlsx"
        export_to_excel(sample_spec, out)

        wb = load_workbook(out)
        ws = wb["Unmapped Variables"]

        # ETHNGRP is the SUPPQUAL candidate in the DM sample spec
        assert ws.cell(row=4, column=2).value == "ETHNGRP"
        assert ws.cell(row=4, column=4).value == "SUPPDM Candidate"

    def test_export_to_excel_creates_parent_dirs(
        self, sample_spec: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """Excel export creates parent directories if they don't exist."""
        out = tmp_path / "nested" / "dirs" / "dm_mapping.xlsx"
        export_to_excel(sample_spec, out)
        assert out.exists()
