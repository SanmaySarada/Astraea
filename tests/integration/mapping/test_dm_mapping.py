"""Integration test for DM domain mapping using real Fakedata and real LLM calls.

This module exercises the full mapping pipeline end-to-end:
    1. Profile real SAS datasets from Fakedata/
    2. Build context with eCRF form metadata
    3. Call Claude for structured mapping proposals
    4. Validate and enrich proposals against SDTM-IG + CT
    5. Export to JSON (and optionally Excel)

Requires ANTHROPIC_API_KEY environment variable to be set.
"""

from __future__ import annotations

import contextlib
import json
import os
from pathlib import Path

import pytest
from rich.console import Console
from rich.table import Table

from astraea.io.sas_reader import read_sas_with_metadata
from astraea.llm.client import AstraeaLLMClient
from astraea.mapping.engine import MappingEngine
from astraea.models.ecrf import ECRFField, ECRFForm
from astraea.models.mapping import (
    ConfidenceLevel,
    DomainMappingSpec,
    MappingPattern,
    StudyMetadata,
)
from astraea.models.profiling import DatasetProfile
from astraea.profiling.profiler import profile_dataset
from astraea.reference import load_ct_reference, load_sdtm_reference

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FAKEDATA_DIR = Path(__file__).resolve().parents[3] / "Fakedata"
STUDY_ID = "PHA022121-C301"

# Cross-domain source datasets that contribute variables to DM
CROSS_DOMAIN_DATASETS = ["ex", "ie", "irt", "ds"]

# All 7 required DM variables per SDTM-IG v3.4
REQUIRED_DM_VARIABLES = {"STUDYID", "DOMAIN", "USUBJID", "SUBJID", "SITEID", "SEX", "COUNTRY"}

# Skip condition: no API key means we cannot run LLM calls
_skip_no_api_key = pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY"),
    reason="ANTHROPIC_API_KEY not set -- skipping integration test",
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _build_demographics_ecrf_form() -> ECRFForm:
    """Build a realistic Demographics eCRF form from known dm.sas7bdat structure.

    Rather than parsing the full 189-page PDF (expensive, slow), we construct
    a representative form from the known variable structure of the Demographics
    dataset. This provides the LLM with sufficient context for mapping.
    """
    fields = [
        ECRFField(
            field_number=1,
            field_name="BRTHYR",
            data_type="YYYY",
            sas_label="Year of Birth",
            coded_values=None,
            field_oid="BRTHYR",
        ),
        ECRFField(
            field_number=2,
            field_name="AGE",
            data_type="3",
            sas_label="Age",
            units="years",
            coded_values=None,
            field_oid="AGE",
        ),
        ECRFField(
            field_number=3,
            field_name="SEX",
            data_type="$1",
            sas_label="Sex",
            coded_values={"M": "Male", "F": "Female"},
            field_oid="SEX",
        ),
        ECRFField(
            field_number=4,
            field_name="ETHNIC",
            data_type="$50",
            sas_label="Ethnicity",
            coded_values={
                "HISPANIC OR LATINO": "Hispanic or Latino",
                "NOT HISPANIC OR LATINO": "Not Hispanic or Latino",
            },
            field_oid="ETHNIC",
        ),
        ECRFField(
            field_number=5,
            field_name="RACEAME",
            data_type="1",
            sas_label="Race - American Indian or Alaska Native",
            coded_values={"1": "Yes", "0": "No"},
            field_oid="RACEAME",
        ),
        ECRFField(
            field_number=6,
            field_name="RACEASI",
            data_type="1",
            sas_label="Race - Asian",
            coded_values={"1": "Yes", "0": "No"},
            field_oid="RACEASI",
        ),
        ECRFField(
            field_number=7,
            field_name="RACEBLA",
            data_type="1",
            sas_label="Race - Black or African American",
            coded_values={"1": "Yes", "0": "No"},
            field_oid="RACEBLA",
        ),
        ECRFField(
            field_number=8,
            field_name="RACENAT",
            data_type="1",
            sas_label="Race - Native Hawaiian or Other Pacific Islander",
            coded_values={"1": "Yes", "0": "No"},
            field_oid="RACENAT",
        ),
        ECRFField(
            field_number=9,
            field_name="RACEWHI",
            data_type="1",
            sas_label="Race - White",
            coded_values={"1": "Yes", "0": "No"},
            field_oid="RACEWHI",
        ),
        ECRFField(
            field_number=10,
            field_name="RACENTRE",
            data_type="1",
            sas_label="Race - Not Reported",
            coded_values={"1": "Yes", "0": "No"},
            field_oid="RACENTRE",
        ),
        ECRFField(
            field_number=11,
            field_name="HEIGHT",
            data_type="5.1",
            sas_label="Height",
            units="cm",
            coded_values=None,
            field_oid="HEIGHT",
        ),
        ECRFField(
            field_number=12,
            field_name="DMCBP",
            data_type="$200",
            sas_label="Childbearing/Reproductive Status",
            coded_values=None,
            field_oid="DMCBP",
        ),
    ]

    return ECRFForm(
        form_name="Demographics",
        fields=fields,
        page_numbers=[15, 16],
    )


def _profile_dataset(filename: str) -> DatasetProfile:
    """Profile a single SAS dataset from Fakedata/."""
    path = FAKEDATA_DIR / filename
    if not path.exists():
        pytest.skip(f"Fakedata file not found: {path}")
    df, meta = read_sas_with_metadata(path)
    return profile_dataset(df, meta)


def _display_mapping_spec(spec: DomainMappingSpec) -> None:
    """Display the mapping spec using Rich for human inspection."""
    console = Console()

    console.print()
    console.rule("[bold blue]DM Mapping Specification[/bold blue]")
    console.print(f"Study: {spec.study_id}")
    console.print(f"Domain: {spec.domain} ({spec.domain_label})")
    console.print(f"Model: {spec.model_used}")
    console.print(f"Timestamp: {spec.mapping_timestamp}")
    console.print(f"Source datasets: {', '.join(spec.source_datasets)}")
    if spec.cross_domain_sources:
        console.print(f"Cross-domain sources: {', '.join(spec.cross_domain_sources)}")
    console.print()

    # Variable mappings table
    table = Table(title=f"Variable Mappings ({spec.total_variables} total)")
    table.add_column("#", style="dim", width=3)
    table.add_column("SDTM Variable", style="bold cyan", width=12)
    table.add_column("Label", width=30)
    table.add_column("Core", width=5)
    table.add_column("Source", width=20)
    table.add_column("Pattern", width=14)
    table.add_column("Conf", width=6)
    table.add_column("Logic", width=50, overflow="fold")

    for i, m in enumerate(spec.variable_mappings, 1):
        # Color code confidence
        if m.confidence_level == ConfidenceLevel.HIGH:
            conf_style = "[green]"
        elif m.confidence_level == ConfidenceLevel.MEDIUM:
            conf_style = "[yellow]"
        else:
            conf_style = "[red]"

        source_str = ""
        if m.source_dataset and m.source_variable:
            source_str = f"{m.source_dataset}.{m.source_variable}"
        elif m.source_variable:
            source_str = m.source_variable
        elif m.assigned_value:
            source_str = f'"{m.assigned_value}"'

        table.add_row(
            str(i),
            m.sdtm_variable,
            m.sdtm_label[:30],
            m.core.value,
            source_str,
            m.mapping_pattern.value,
            f"{conf_style}{m.confidence:.2f}[/]",
            m.mapping_logic[:50],
        )

    console.print(table)

    # Summary
    console.print()
    console.print("[bold]Summary:[/bold]")
    console.print(f"  Total variables: {spec.total_variables}")
    console.print(f"  Required mapped: {spec.required_mapped}")
    console.print(f"  Expected mapped: {spec.expected_mapped}")
    console.print(
        f"  Confidence: [green]{spec.high_confidence_count} HIGH[/green] | "
        f"[yellow]{spec.medium_confidence_count} MEDIUM[/yellow] | "
        f"[red]{spec.low_confidence_count} LOW[/red]"
    )

    if spec.unmapped_source_variables:
        unmapped_str = ", ".join(spec.unmapped_source_variables)
        console.print(f"\n[dim]Unmapped source variables: {unmapped_str}[/dim]")
    if spec.suppqual_candidates:
        console.print(f"[dim]SUPPDM candidates: {', '.join(spec.suppqual_candidates)}[/dim]")

    console.rule()


# ---------------------------------------------------------------------------
# Module-scoped fixture: run mapping once, reuse across tests
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def dm_mapping_result() -> DomainMappingSpec:
    """Run the full DM mapping pipeline once and cache the result.

    This is the only fixture that makes a real LLM call. All tests in this
    module share the same result to avoid redundant API calls.
    """
    if not os.environ.get("ANTHROPIC_API_KEY"):
        pytest.skip("ANTHROPIC_API_KEY not set")

    # Profile primary dataset
    dm_profile = _profile_dataset("dm.sas7bdat")

    # Profile cross-domain datasets
    cross_domain_profiles: dict[str, DatasetProfile] = {}
    for ds_name in CROSS_DOMAIN_DATASETS:
        filename = f"{ds_name}.sas7bdat"
        with contextlib.suppress(Exception):
            cross_domain_profiles[ds_name] = _profile_dataset(filename)

    # Build eCRF form
    ecrf_form = _build_demographics_ecrf_form()

    # Initialize references
    sdtm_ref = load_sdtm_reference()
    ct_ref = load_ct_reference()

    # Initialize LLM client and engine
    llm_client = AstraeaLLMClient()
    engine = MappingEngine(llm_client, sdtm_ref, ct_ref)

    # Study metadata
    study_metadata = StudyMetadata(
        study_id=STUDY_ID,
        site_id_variable="SiteNumber",
        subject_id_variable="Subject",
    )

    # Run the mapping
    spec = engine.map_domain(
        domain="DM",
        source_profiles=[dm_profile],
        ecrf_forms=[ecrf_form],
        study_metadata=study_metadata,
        cross_domain_profiles=cross_domain_profiles,
    )

    # Display for human inspection (visible with pytest -s)
    _display_mapping_spec(spec)

    return spec


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.integration
@_skip_no_api_key
class TestDMMappingEndToEnd:
    """End-to-end tests for DM domain mapping on real Fakedata."""

    def test_domain_is_dm(self, dm_mapping_result: DomainMappingSpec) -> None:
        """Mapping result targets the DM domain."""
        assert dm_mapping_result.domain == "DM"

    def test_sufficient_variable_count(self, dm_mapping_result: DomainMappingSpec) -> None:
        """At least 15 variables should be mapped (DM has 27 possible)."""
        assert dm_mapping_result.total_variables >= 15, (
            f"Expected at least 15 mapped variables, got {dm_mapping_result.total_variables}"
        )

    def test_all_required_variables_mapped(self, dm_mapping_result: DomainMappingSpec) -> None:
        """All 7 Required DM variables must have mappings."""
        mapped_vars = {m.sdtm_variable for m in dm_mapping_result.variable_mappings}
        missing = REQUIRED_DM_VARIABLES - mapped_vars
        assert not missing, (
            f"Missing required DM variables: {missing}. Mapped variables: {sorted(mapped_vars)}"
        )

    def test_studyid_is_assign(self, dm_mapping_result: DomainMappingSpec) -> None:
        """STUDYID should be assigned as a constant value."""
        studyid_mappings = [
            m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "STUDYID"
        ]
        assert len(studyid_mappings) == 1
        assert studyid_mappings[0].mapping_pattern == MappingPattern.ASSIGN

    def test_domain_is_assign_dm(self, dm_mapping_result: DomainMappingSpec) -> None:
        """DOMAIN should be assigned as constant 'DM'."""
        domain_mappings = [
            m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "DOMAIN"
        ]
        assert len(domain_mappings) == 1
        m = domain_mappings[0]
        assert m.mapping_pattern == MappingPattern.ASSIGN
        assert m.assigned_value is not None
        assert m.assigned_value.upper() == "DM"

    def test_age_mapping_exists(self, dm_mapping_result: DomainMappingSpec) -> None:
        """AGE should be mapped (direct from source or derived)."""
        age_mappings = [m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "AGE"]
        assert len(age_mappings) >= 1, "AGE mapping not found in DM spec"

    def test_race_mapping_exists(self, dm_mapping_result: DomainMappingSpec) -> None:
        """RACE should be mapped, likely as COMBINE from checkbox columns."""
        race_mappings = [
            m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "RACE"
        ]
        assert len(race_mappings) >= 1, "RACE mapping not found in DM spec"

    def test_cross_domain_sources_present(self, dm_mapping_result: DomainMappingSpec) -> None:
        """At least one cross-domain source dataset should be referenced.

        For DM, we expect references to ex.sas7bdat (for RFSTDTC), irt.sas7bdat
        (for COUNTRY), or others.
        """
        # Check either cross_domain_sources list or variable source_dataset references
        all_source_datasets = {
            m.source_dataset
            for m in dm_mapping_result.variable_mappings
            if m.source_dataset is not None
        }
        cross_domain_files = {f"{name}.sas7bdat" for name in CROSS_DOMAIN_DATASETS}
        has_cross_domain = bool(all_source_datasets & cross_domain_files) or bool(
            dm_mapping_result.cross_domain_sources
        )
        assert has_cross_domain, (
            f"No cross-domain sources found. Source datasets: {all_source_datasets}, "
            f"cross_domain_sources: {dm_mapping_result.cross_domain_sources}"
        )

    def test_high_confidence_exists(self, dm_mapping_result: DomainMappingSpec) -> None:
        """At least some mappings should have HIGH confidence."""
        assert dm_mapping_result.high_confidence_count > 0, (
            "No HIGH confidence mappings found. Distribution: "
            f"HIGH={dm_mapping_result.high_confidence_count}, "
            f"MEDIUM={dm_mapping_result.medium_confidence_count}, "
            f"LOW={dm_mapping_result.low_confidence_count}"
        )

    def test_required_variables_have_reasonable_confidence(
        self, dm_mapping_result: DomainMappingSpec
    ) -> None:
        """No Required variable should have confidence below 0.5."""
        for m in dm_mapping_result.variable_mappings:
            if m.sdtm_variable in REQUIRED_DM_VARIABLES:
                assert m.confidence >= 0.5, (
                    f"Required variable {m.sdtm_variable} has low confidence: "
                    f"{m.confidence:.2f} ({m.confidence_rationale})"
                )


@pytest.mark.integration
@_skip_no_api_key
class TestDMExportRoundtrip:
    """Test JSON export and round-trip validation."""

    def test_json_export_roundtrip(
        self, dm_mapping_result: DomainMappingSpec, tmp_path: Path
    ) -> None:
        """Export to JSON, read back, and validate it round-trips to DomainMappingSpec."""
        json_path = tmp_path / "dm_mapping.json"
        json_path.write_text(dm_mapping_result.model_dump_json(indent=2))

        assert json_path.exists()
        assert json_path.stat().st_size > 0

        # Read back and validate
        raw = json.loads(json_path.read_text())
        roundtripped = DomainMappingSpec.model_validate(raw)

        assert roundtripped.domain == dm_mapping_result.domain
        assert roundtripped.total_variables == dm_mapping_result.total_variables
        assert len(roundtripped.variable_mappings) == len(dm_mapping_result.variable_mappings)

    def test_excel_export(self, dm_mapping_result: DomainMappingSpec, tmp_path: Path) -> None:
        """Export to Excel and verify the workbook has expected structure.

        Uses a minimal in-test Excel writer since the exporters module
        (plan 03-04) has not been built yet. Validates the spec data can
        be serialized to a 3-sheet workbook.
        """
        from openpyxl import Workbook, load_workbook

        xlsx_path = tmp_path / "dm_mapping.xlsx"
        spec = dm_mapping_result

        # Write workbook
        wb = Workbook()

        # Sheet 1: Mapping Spec
        ws1 = wb.active
        ws1.title = "Mapping Spec"  # type: ignore[union-attr]
        headers = [
            "#",
            "SDTM Variable",
            "SDTM Label",
            "Type",
            "Core",
            "Source Dataset",
            "Source Variable",
            "Pattern",
            "Mapping Logic",
            "Confidence",
            "Confidence Level",
        ]
        ws1.append(headers)  # type: ignore[union-attr]
        for i, m in enumerate(spec.variable_mappings, 1):
            ws1.append(
                [  # type: ignore[union-attr]
                    i,
                    m.sdtm_variable,
                    m.sdtm_label,
                    m.sdtm_data_type,
                    m.core.value,
                    m.source_dataset or "",
                    m.source_variable or "",
                    m.mapping_pattern.value,
                    m.mapping_logic,
                    m.confidence,
                    m.confidence_level.value,
                ]
            )

        # Sheet 2: Unmapped Variables
        ws2 = wb.create_sheet("Unmapped Variables")
        ws2.append(["Source Variable", "Disposition"])
        for var in spec.unmapped_source_variables:
            ws2.append([var, "Unmapped"])
        for var in spec.suppqual_candidates:
            ws2.append([var, "SUPPDM Candidate"])

        # Sheet 3: Summary
        ws3 = wb.create_sheet("Summary")
        summary_rows = [
            ("Domain", spec.domain),
            ("Domain Label", spec.domain_label),
            ("Study ID", spec.study_id),
            ("Timestamp", spec.mapping_timestamp),
            ("Model", spec.model_used),
            ("Total Variables", spec.total_variables),
            ("Required Mapped", spec.required_mapped),
            ("Expected Mapped", spec.expected_mapped),
            ("HIGH Confidence", spec.high_confidence_count),
            ("MEDIUM Confidence", spec.medium_confidence_count),
            ("LOW Confidence", spec.low_confidence_count),
        ]
        for key, val in summary_rows:
            ws3.append([key, val])

        wb.save(xlsx_path)

        # Verify
        assert xlsx_path.exists()
        wb2 = load_workbook(xlsx_path)
        assert len(wb2.sheetnames) == 3
        assert "Mapping Spec" in wb2.sheetnames
        assert "Unmapped Variables" in wb2.sheetnames
        assert "Summary" in wb2.sheetnames

        # Verify row counts (header + data rows)
        ws_spec = wb2["Mapping Spec"]
        data_rows = ws_spec.max_row - 1  # subtract header
        assert data_rows == spec.total_variables, (
            f"Expected {spec.total_variables} data rows, got {data_rows}"
        )


@pytest.mark.integration
@_skip_no_api_key
class TestDMCTValidation:
    """Validate controlled terminology references in the mapping output."""

    def test_sex_references_correct_codelist(self, dm_mapping_result: DomainMappingSpec) -> None:
        """SEX mapping should reference codelist C66731 (Sex)."""
        sex_mappings = [m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "SEX"]
        assert len(sex_mappings) >= 1
        sex = sex_mappings[0]
        # The LLM may or may not include codelist_code; if it does, it should be C66731
        if sex.codelist_code is not None:
            assert sex.codelist_code == "C66731", (
                f"SEX codelist should be C66731, got {sex.codelist_code}"
            )

    def test_ethnic_references_correct_codelist(self, dm_mapping_result: DomainMappingSpec) -> None:
        """ETHNIC mapping should reference codelist C66790 (Ethnicity)."""
        ethnic_mappings = [
            m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "ETHNIC"
        ]
        assert len(ethnic_mappings) >= 1
        ethnic = ethnic_mappings[0]
        if ethnic.codelist_code is not None:
            assert ethnic.codelist_code == "C66790", (
                f"ETHNIC codelist should be C66790, got {ethnic.codelist_code}"
            )

    def test_race_has_reasonable_codelist(self, dm_mapping_result: DomainMappingSpec) -> None:
        """RACE mapping should reference C74457 or have reasonable confidence."""
        race_mappings = [
            m for m in dm_mapping_result.variable_mappings if m.sdtm_variable == "RACE"
        ]
        assert len(race_mappings) >= 1
        race = race_mappings[0]
        # Either references the correct codelist or has reasonable confidence
        if race.codelist_code is not None:
            assert race.codelist_code == "C74457", (
                f"RACE codelist should be C74457, got {race.codelist_code}"
            )
        else:
            # If no codelist provided, confidence should still be reasonable
            assert race.confidence >= 0.5, (
                f"RACE without codelist should have confidence >= 0.5, got {race.confidence:.2f}"
            )
